"""
Conversion Metadata Tracker and Excel Dashboard Generator

Tracks conversion metadata and generates an Excel dashboard with:
- ISBN, Publisher, Conversion Date/Time
- Status (Progress, Success, Failure)
- Type (PDF, ePub)
- Template (Single Column, Double Column)
- Image counts (vector, raster)
- Table counts

Appends to existing Excel file or creates new one.
"""

from dataclasses import dataclass, field
from typing import Optional, List, Dict
from pathlib import Path
from datetime import datetime
import logging
from enum import Enum
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ConversionStatus(Enum):
    """Conversion status states"""
    IN_PROGRESS = "In Progress"
    SUCCESS = "Success"
    FAILURE = "Failure"
    PARTIAL = "Partial Success"


class ConversionType(Enum):
    """Source file type"""
    PDF = "PDF"
    EPUB = "ePub"
    DOCX = "DOCX"


class TemplateType(Enum):
    """Document template/layout type"""
    SINGLE_COLUMN = "Single Column"
    DOUBLE_COLUMN = "Double Column"
    MIXED = "Mixed"
    UNKNOWN = "Unknown"


@dataclass
class ConversionMetadata:
    """Metadata for a single conversion job"""
    # Identifiers
    filename: str
    isbn: Optional[str] = None
    publisher: Optional[str] = None

    # Timing
    start_time: datetime = field(default_factory=datetime.now)
    end_time: Optional[datetime] = None

    # Status
    status: ConversionStatus = ConversionStatus.IN_PROGRESS
    progress_percent: int = 0
    error_message: Optional[str] = None

    # Source info
    conversion_type: ConversionType = ConversionType.PDF
    template_type: TemplateType = TemplateType.UNKNOWN

    # Statistics
    num_chapters: int = 0
    num_pages: int = 0
    num_vector_images: int = 0
    num_raster_images: int = 0
    num_tables: int = 0
    num_equations: int = 0

    # Output
    output_path: Optional[str] = None
    output_size_mb: Optional[float] = None

    # Additional metadata
    title: Optional[str] = None
    authors: List[str] = field(default_factory=list)
    language: Optional[str] = None

    def duration_seconds(self) -> Optional[int]:
        """Calculate conversion duration in seconds"""
        if self.end_time:
            return int((self.end_time - self.start_time).total_seconds())
        return None

    def to_row(self) -> List:
        """Convert to Excel row format"""
        duration = self.duration_seconds()
        duration_str = f"{duration}s" if duration else "N/A"

        authors_str = ", ".join(self.authors) if self.authors else ""

        return [
            self.filename,
            self.isbn or "",
            self.title or "",
            self.publisher or "",
            authors_str,
            self.start_time.strftime("%Y-%m-%d %H:%M:%S"),
            self.end_time.strftime("%Y-%m-%d %H:%M:%S") if self.end_time else "",
            duration_str,
            self.status.value,
            self.progress_percent,
            self.conversion_type.value,
            self.template_type.value,
            self.num_chapters,
            self.num_pages,
            self.num_vector_images,
            self.num_raster_images,
            self.num_vector_images + self.num_raster_images,  # Total images
            self.num_tables,
            self.num_equations,
            self.output_path or "",
            f"{self.output_size_mb:.2f}" if self.output_size_mb else "",
            self.error_message or "",
        ]


class ConversionTracker:
    """
    Manages conversion tracking and Excel dashboard generation.
    """

    EXCEL_FILENAME = "conversion_dashboard.xlsx"
    SHEET_NAME = "Conversions"

    HEADERS = [
        "Filename",
        "ISBN",
        "Title",
        "Publisher",
        "Authors",
        "Start Time",
        "End Time",
        "Duration",
        "Status",
        "Progress %",
        "Type",
        "Template",
        "# Chapters",
        "# Pages",
        "# Vector Images",
        "# Raster Images",
        "# Total Images",
        "# Tables",
        "# Equations",
        "Output Path",
        "Output Size (MB)",
        "Error Message",
    ]

    def __init__(self, output_dir: Path):
        """
        Initialize tracker.

        Args:
            output_dir: Directory where Excel file will be created/updated
        """
        self.output_dir = Path(output_dir)
        self.excel_path = self.output_dir / self.EXCEL_FILENAME
        self.current_metadata: Optional[ConversionMetadata] = None

    def start_conversion(self,
                        filename: str,
                        conversion_type: ConversionType,
                        **kwargs) -> ConversionMetadata:
        """
        Start tracking a new conversion.

        Args:
            filename: Input filename
            conversion_type: PDF or ePub
            **kwargs: Additional metadata fields

        Returns:
            ConversionMetadata object
        """
        self.current_metadata = ConversionMetadata(
            filename=filename,
            conversion_type=conversion_type,
            start_time=datetime.now(),
            **kwargs
        )

        logger.info(f"Started tracking conversion: {filename}")
        return self.current_metadata

    def update_progress(self, progress_percent: int, status: Optional[ConversionStatus] = None) -> None:
        """Update conversion progress"""
        if self.current_metadata:
            self.current_metadata.progress_percent = min(100, max(0, progress_percent))
            if status:
                self.current_metadata.status = status

            # Write intermediate update to Excel
            self._save_to_excel()

    def complete_conversion(self,
                           status: ConversionStatus,
                           error_message: Optional[str] = None,
                           **kwargs) -> None:
        """
        Mark conversion as complete.

        Args:
            status: Final status (SUCCESS or FAILURE)
            error_message: Error message if failed
            **kwargs: Additional metadata to update
        """
        if not self.current_metadata:
            logger.warning("complete_conversion called without active conversion")
            return

        self.current_metadata.end_time = datetime.now()
        self.current_metadata.status = status
        self.current_metadata.progress_percent = 100 if status == ConversionStatus.SUCCESS else self.current_metadata.progress_percent

        if error_message:
            self.current_metadata.error_message = error_message

        # Update any additional fields
        for key, value in kwargs.items():
            if hasattr(self.current_metadata, key):
                setattr(self.current_metadata, key, value)

        # Save final state
        self._save_to_excel()

        logger.info(f"Completed tracking conversion: {self.current_metadata.filename} - {status.value}")

    def _save_to_excel(self) -> None:
        """Save current metadata to Excel file (append or update)"""
        if not self.current_metadata:
            return

        try:
            # Load or create workbook
            if self.excel_path.exists():
                wb = openpyxl.load_workbook(self.excel_path)
                if self.SHEET_NAME in wb.sheetnames:
                    ws = wb[self.SHEET_NAME]
                    is_new_sheet = False
                else:
                    ws = wb.create_sheet(self.SHEET_NAME)
                    is_new_sheet = True
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = self.SHEET_NAME
                is_new_sheet = True

            # Add headers if new sheet
            if is_new_sheet or ws.max_row == 0:
                ws.append(self.HEADERS)
                self._format_header_row(ws)

            # Check if this conversion already exists (update mode)
            row_to_update = None
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row_idx, 1).value == self.current_metadata.filename and \
                   ws.cell(row_idx, 6).value == self.current_metadata.start_time.strftime("%Y-%m-%d %H:%M:%S"):
                    row_to_update = row_idx
                    break

            # Get data row
            data_row = self.current_metadata.to_row()

            if row_to_update:
                # Update existing row
                for col_idx, value in enumerate(data_row, start=1):
                    ws.cell(row_to_update, col_idx).value = value
                logger.debug(f"Updated row {row_to_update} in Excel")
            else:
                # Append new row
                ws.append(data_row)
                row_to_update = ws.max_row
                logger.debug(f"Added new row {row_to_update} in Excel")

            # Format the data row
            self._format_data_row(ws, row_to_update, self.current_metadata.status)

            # Auto-size columns
            self._auto_size_columns(ws)

            # Save workbook
            wb.save(self.excel_path)
            logger.info(f"Saved conversion tracking to {self.excel_path}")

        except Exception as e:
            logger.error(f"Failed to save to Excel: {e}", exc_info=True)

    def _format_header_row(self, ws) -> None:
        """Format the header row with colors and styles"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment

    def _format_data_row(self, ws, row_idx: int, status: ConversionStatus) -> None:
        """Format a data row based on status"""
        # Status column color coding
        status_col = 9  # "Status" column
        progress_col = 10  # "Progress %" column

        if status == ConversionStatus.SUCCESS:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            font = Font(color="006100")
        elif status == ConversionStatus.FAILURE:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            font = Font(color="9C0006")
        elif status == ConversionStatus.IN_PROGRESS:
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            font = Font(color="9C6500")
        else:
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            font = Font(color="000000")

        ws.cell(row_idx, status_col).fill = fill
        ws.cell(row_idx, status_col).font = font

        # Add progress bar effect in Progress % column
        progress_val = ws.cell(row_idx, progress_col).value
        if isinstance(progress_val, int):
            # Color-code progress
            if progress_val == 100:
                ws.cell(row_idx, progress_col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif progress_val > 0:
                ws.cell(row_idx, progress_col).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def _auto_size_columns(self, ws) -> None:
        """Auto-size columns based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def get_statistics(self) -> Dict:
        """Get statistics from all conversions in the Excel file"""
        if not self.excel_path.exists():
            return {}

        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb[self.SHEET_NAME]

            stats = {
                'total_conversions': ws.max_row - 1,  # Exclude header
                'successful': 0,
                'failed': 0,
                'in_progress': 0,
                'total_images': 0,
                'total_tables': 0,
                'pdf_conversions': 0,
                'epub_conversions': 0,
            }

            for row_idx in range(2, ws.max_row + 1):
                status = ws.cell(row_idx, 9).value
                conv_type = ws.cell(row_idx, 11).value
                total_images = ws.cell(row_idx, 17).value
                total_tables = ws.cell(row_idx, 18).value

                if status == "Success":
                    stats['successful'] += 1
                elif status == "Failure":
                    stats['failed'] += 1
                elif status == "In Progress":
                    stats['in_progress'] += 1

                if conv_type == "PDF":
                    stats['pdf_conversions'] += 1
                elif conv_type == "ePub":
                    stats['epub_conversions'] += 1

                if isinstance(total_images, int):
                    stats['total_images'] += total_images
                if isinstance(total_tables, int):
                    stats['total_tables'] += total_tables

            return stats

        except Exception as e:
            logger.error(f"Failed to get statistics: {e}")
            return {}


# Convenience function for single-file tracking
def track_conversion(output_dir: Path,
                    filename: str,
                    conversion_type: ConversionType,
                    **metadata) -> ConversionTracker:
    """
    Create and start a conversion tracker.

    Args:
        output_dir: Output directory for Excel file
        filename: Input filename
        conversion_type: PDF or ePub
        **metadata: Additional metadata fields

    Returns:
        ConversionTracker instance
    """
    tracker = ConversionTracker(output_dir)
    tracker.start_conversion(filename, conversion_type, **metadata)
    return tracker
