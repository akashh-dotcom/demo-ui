#!/usr/bin/env python3
"""
Validation Report Generator

Generates Excel reports with validation errors in plain English.
Includes XML filename, line number, error type, and detailed error description.
"""

import os
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional

try:
    from lxml import etree
    LXML_AVAILABLE = True
except ImportError:
    LXML_AVAILABLE = False
    etree = None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@dataclass
class ValidationError:
    """Represents a single validation error"""
    xml_file: str
    line_number: Optional[int]
    column_number: Optional[int]
    error_type: str
    error_description: str
    severity: str  # 'Error', 'Warning', 'Info'

    def to_dict(self):
        """Convert to dictionary for export"""
        return {
            'xml_file': self.xml_file,
            'line_number': self.line_number or 'N/A',
            'column_number': self.column_number or 'N/A',
            'error_type': self.error_type,
            'error_description': self.error_description,
            'severity': self.severity
        }


@dataclass
class VerificationItem:
    """Represents an item that requires manual content verification"""
    xml_file: str
    line_number: Optional[int]
    fix_type: str  # Type of automated fix applied
    fix_description: str  # What was changed
    verification_reason: str  # Why manual verification is needed
    suggestion: str  # Suggested action for manual verification

    def to_dict(self):
        """Convert to dictionary for export"""
        return {
            'xml_file': self.xml_file,
            'line_number': self.line_number or 'N/A',
            'fix_type': self.fix_type,
            'fix_description': self.fix_description,
            'verification_reason': self.verification_reason,
            'suggestion': self.suggestion
        }


class ValidationReportGenerator:
    """Generates validation reports in Excel format"""

    def __init__(self):
        self.errors: List[ValidationError] = []
        self.verification_items: List[VerificationItem] = []

    def clear(self):
        """Clear all errors and verification items"""
        self.errors = []
        self.verification_items = []

    def add_error(self, error: ValidationError):
        """Add a validation error to the report"""
        self.errors.append(error)

    def add_verification_item(self, item: VerificationItem):
        """Add a verification item to the report"""
        self.verification_items.append(item)

    def parse_dtd_errors(self, xml_filename: str, error_log) -> List[ValidationError]:
        """
        Parse lxml DTD error log and extract structured errors.

        Args:
            xml_filename: Name of the XML file being validated
            error_log: lxml DTD error log object

        Returns:
            List of ValidationError objects
        """
        errors = []

        for error in error_log:
            # Parse line and column from error
            line_num = error.line if hasattr(error, 'line') else None
            col_num = error.column if hasattr(error, 'column') else None

            # Get error message
            message = str(error.message) if hasattr(error, 'message') else str(error)

            # Categorize error type from message
            error_type = self._categorize_dtd_error(message)

            # Make description more readable
            description = self._make_description_readable(message)

            errors.append(ValidationError(
                xml_file=xml_filename,
                line_number=line_num,
                column_number=col_num,
                error_type=error_type,
                error_description=description,
                severity='Error'
            ))

        return errors

    def _categorize_dtd_error(self, message: str) -> str:
        """Categorize DTD error based on message content"""
        message_lower = message.lower()

        if 'no declaration' in message_lower or 'not declared' in message_lower:
            return 'Undeclared Element'
        elif 'does not follow' in message_lower or 'content model' in message_lower:
            return 'Invalid Content Model'
        elif 'not allowed' in message_lower or 'unexpected' in message_lower:
            return 'Invalid Element'
        elif 'required attribute' in message_lower or 'missing' in message_lower:
            return 'Missing Attribute'
        elif 'invalid attribute' in message_lower:
            return 'Invalid Attribute'
        elif 'id' in message_lower and 'already defined' in message_lower:
            return 'Duplicate ID'
        elif 'entity' in message_lower:
            return 'Entity Error'
        elif 'empty' in message_lower:
            return 'Empty Element Error'
        else:
            return 'DTD Validation Error'

    def _make_description_readable(self, message: str) -> str:
        """Convert technical DTD error messages to plain English"""

        # Common patterns to replace
        replacements = {
            r'Element (\w+): ': r'The <\1> element ',
            r'No declaration for element (\w+)': r'The element <\1> is not defined in the DTD',
            r'No declaration for attribute (\w+) of element (\w+)':
                r'The attribute "\1" is not allowed in the <\2> element',
            r'does not follow the DTD': r'does not match what the DTD expects',
            r'content model': r'allowed content',
            r'Expecting \(([^)]+)\)': r'Expected: \1',
            r'got \(([^)]+)\)': r'but found: \1',
        }

        readable = message
        for pattern, replacement in replacements.items():
            readable = re.sub(pattern, replacement, readable)

        return readable

    def add_reference_errors(self, xml_filename: str, reference_errors: List[str]):
        """
        Add reference validation errors from ReferenceMapper.

        Args:
            xml_filename: Name of the XML file or package
            reference_errors: List of error messages from reference validation
        """
        for error_msg in reference_errors:
            # Try to extract more details from the error message
            if 'Final resource not found' in error_msg:
                error_type = 'Missing Resource'
            elif 'Unresolved link' in error_msg:
                error_type = 'Broken Link'
            elif 'has no final name' in error_msg:
                error_type = 'Missing Resource Name'
            else:
                error_type = 'Reference Error'

            self.errors.append(ValidationError(
                xml_file=xml_filename,
                line_number=None,
                column_number=None,
                error_type=error_type,
                error_description=error_msg,
                severity='Warning' if 'Unresolved' in error_msg else 'Error'
            ))

    def add_xml_syntax_error(self, xml_filename: str, syntax_error):
        """
        Add XML syntax error from lxml parsing.

        Args:
            xml_filename: Name of the XML file
            syntax_error: XMLSyntaxError exception
        """
        self.errors.append(ValidationError(
            xml_file=xml_filename,
            line_number=syntax_error.lineno if hasattr(syntax_error, 'lineno') else None,
            column_number=None,
            error_type='XML Syntax Error',
            error_description=str(syntax_error),
            severity='Error'
        ))

    def add_general_error(self, xml_filename: str, error_type: str,
                         description: str, severity: str = 'Error'):
        """
        Add a general validation error.

        Args:
            xml_filename: Name of the XML file
            error_type: Type of error
            description: Error description
            severity: Error severity ('Error', 'Warning', 'Info')
        """
        self.errors.append(ValidationError(
            xml_file=xml_filename,
            line_number=None,
            column_number=None,
            error_type=error_type,
            error_description=description,
            severity=severity
        ))

    def generate_excel_report(self, output_path: Path, book_title: str = "Unknown"):
        """
        Generate an Excel validation report.

        Args:
            output_path: Path where to save the Excel file
            book_title: Title of the book being validated
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Errors"

        # Create summary sheet
        self._create_summary_sheet(wb, book_title)

        # Create manual verification sheet if we have verification items
        if self.verification_items:
            self._create_verification_sheet(wb)

        # Switch to details sheet
        ws = wb["Validation Errors"]

        # Define headers
        headers = [
            "Error #",
            "Severity",
            "XML File",
            "Line Number",
            "Column",
            "Error Type",
            "Description"
        ]

        # Style definitions
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        info_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Write data rows
        for idx, error in enumerate(self.errors, 1):
            row_num = idx + 1

            # Choose fill color based on severity
            if error.severity == 'Error':
                row_fill = error_fill
            elif error.severity == 'Warning':
                row_fill = warning_fill
            else:
                row_fill = info_fill

            # Error number
            cell = ws.cell(row=row_num, column=1, value=idx)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            cell.fill = row_fill

            # Severity
            cell = ws.cell(row=row_num, column=2, value=error.severity)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            cell.fill = row_fill
            if error.severity == 'Error':
                cell.font = Font(bold=True, color="9C0006")
            elif error.severity == 'Warning':
                cell.font = Font(bold=True, color="9C6500")

            # XML File
            cell = ws.cell(row=row_num, column=3, value=error.xml_file)
            cell.border = border
            cell.fill = row_fill

            # Line Number
            line_value = error.line_number if error.line_number else "N/A"
            cell = ws.cell(row=row_num, column=4, value=line_value)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            cell.fill = row_fill

            # Column
            col_value = error.column_number if error.column_number else "N/A"
            cell = ws.cell(row=row_num, column=5, value=col_value)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            cell.fill = row_fill

            # Error Type
            cell = ws.cell(row=row_num, column=6, value=error.error_type)
            cell.border = border
            cell.fill = row_fill

            # Description
            cell = ws.cell(row=row_num, column=7, value=error.error_description)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border
            cell.fill = row_fill

        # Adjust column widths
        ws.column_dimensions['A'].width = 10  # Error #
        ws.column_dimensions['B'].width = 12  # Severity
        ws.column_dimensions['C'].width = 25  # XML File
        ws.column_dimensions['D'].width = 12  # Line Number
        ws.column_dimensions['E'].width = 10  # Column
        ws.column_dimensions['F'].width = 25  # Error Type
        ws.column_dimensions['G'].width = 60  # Description

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save workbook with retry logic for Windows/OneDrive permission issues
        self._save_workbook_with_retry(wb, output_path)

    def _create_summary_sheet(self, wb: Workbook, book_title: str):
        """Create a summary sheet with overview statistics"""
        ws = wb.create_sheet("Summary", 0)

        # Count errors by type and severity
        total_errors = len(self.errors)
        errors_by_severity = {'Error': 0, 'Warning': 0, 'Info': 0}
        errors_by_type = {}

        for error in self.errors:
            errors_by_severity[error.severity] = errors_by_severity.get(error.severity, 0) + 1
            errors_by_type[error.error_type] = errors_by_type.get(error.error_type, 0) + 1

        # Title styling
        title_font = Font(bold=True, size=16, color="366092")
        header_font = Font(bold=True, size=12)

        # Report title
        ws['A1'] = 'VALIDATION REPORT'
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        # Metadata
        row = 3
        ws[f'A{row}'] = 'Book Title:'
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = book_title

        row += 1
        ws[f'A{row}'] = 'Report Generated:'
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        row += 2
        ws[f'A{row}'] = 'SUMMARY STATISTICS'
        ws[f'A{row}'].font = Font(bold=True, size=14, color="366092")

        row += 2
        ws[f'A{row}'] = 'Total Errors:'
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = total_errors
        ws[f'B{row}'].font = Font(bold=True, size=12)

        # Errors by severity
        row += 2
        ws[f'A{row}'] = 'ERRORS BY SEVERITY'
        ws[f'A{row}'].font = Font(bold=True, size=12)

        for severity, count in sorted(errors_by_severity.items()):
            row += 1
            ws[f'A{row}'] = f'{severity}:'
            ws[f'B{row}'] = count

            if severity == 'Error' and count > 0:
                ws[f'B{row}'].font = Font(bold=True, color="9C0006")
            elif severity == 'Warning' and count > 0:
                ws[f'B{row}'].font = Font(bold=True, color="9C6500")

        # Errors by type (top 10)
        row += 2
        ws[f'A{row}'] = 'TOP ERROR TYPES'
        ws[f'A{row}'].font = Font(bold=True, size=12)

        sorted_types = sorted(errors_by_type.items(), key=lambda x: x[1], reverse=True)[:10]
        for error_type, count in sorted_types:
            row += 1
            ws[f'A{row}'] = f'{error_type}:'
            ws[f'B{row}'] = count

        # Manual verification items
        if self.verification_items:
            row += 2
            ws[f'A{row}'] = 'MANUAL VERIFICATION REQUIRED'
            ws[f'A{row}'].font = Font(bold=True, size=12)

            row += 1
            ws[f'A{row}'] = 'Items requiring review:'
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'] = len(self.verification_items)
            ws[f'B{row}'].font = Font(bold=True, size=12, color="9C6500")
            ws[f'B{row}'].fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

            row += 1
            ws[f'A{row}'] = 'See "Manual Verification" sheet for details'
            ws[f'A{row}'].font = Font(italic=True, color="666666")
            ws.merge_cells(f'A{row}:D{row}')

        # Validation status
        row += 2
        ws[f'A{row}'] = 'VALIDATION STATUS:'
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        if total_errors == 0:
            ws[f'A{row}'] = '✓ PASSED - No errors found'
            ws[f'A{row}'].font = Font(bold=True, size=12, color="006100")
            ws[f'A{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws[f'A{row}'] = f'✗ FAILED - {total_errors} error(s) found'
            ws[f'A{row}'].font = Font(bold=True, size=12, color="9C0006")
            ws[f'A{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        ws.merge_cells(f'A{row}:D{row}')

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_verification_sheet(self, wb: Workbook):
        """Create a sheet for items requiring manual verification"""
        ws = wb.create_sheet("Manual Verification")

        # Define headers
        headers = [
            "Item #",
            "XML File",
            "Line Number",
            "Fix Type",
            "What Was Changed",
            "Why Verification Needed",
            "Suggested Action"
        ]

        # Style definitions
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        verification_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Write data rows
        for idx, item in enumerate(self.verification_items, 1):
            row_num = idx + 1

            # Item number
            cell = ws.cell(row=row_num, column=1, value=idx)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            cell.fill = verification_fill

            # XML File
            cell = ws.cell(row=row_num, column=2, value=item.xml_file)
            cell.border = border
            cell.fill = verification_fill

            # Line Number
            line_value = item.line_number if item.line_number else "N/A"
            cell = ws.cell(row=row_num, column=3, value=line_value)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            cell.fill = verification_fill

            # Fix Type
            cell = ws.cell(row=row_num, column=4, value=item.fix_type)
            cell.border = border
            cell.fill = verification_fill

            # What Was Changed
            cell = ws.cell(row=row_num, column=5, value=item.fix_description)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border
            cell.fill = verification_fill

            # Why Verification Needed
            cell = ws.cell(row=row_num, column=6, value=item.verification_reason)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border
            cell.fill = verification_fill

            # Suggested Action
            cell = ws.cell(row=row_num, column=7, value=item.suggestion)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border
            cell.fill = verification_fill

        # Adjust column widths
        ws.column_dimensions['A'].width = 10  # Item #
        ws.column_dimensions['B'].width = 20  # XML File
        ws.column_dimensions['C'].width = 12  # Line Number
        ws.column_dimensions['D'].width = 25  # Fix Type
        ws.column_dimensions['E'].width = 40  # What Was Changed
        ws.column_dimensions['F'].width = 40  # Why Verification Needed
        ws.column_dimensions['G'].width = 40  # Suggested Action

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _save_workbook_with_retry(self, wb: Workbook, output_path: Path, max_retries: int = 3):
        """
        Save workbook with retry logic for Windows/OneDrive permission issues.

        Args:
            wb: Workbook to save
            output_path: Path to save to
            max_retries: Maximum number of retry attempts

        Raises:
            PermissionError: If file cannot be saved after retries
        """
        last_error = None

        for attempt in range(max_retries):
            try:
                # Try to remove existing file if it exists
                if output_path.exists():
                    try:
                        os.remove(output_path)
                        time.sleep(0.1)  # Brief pause after deletion
                    except PermissionError:
                        # File might be open, continue and try to overwrite
                        pass

                # Try to save
                wb.save(output_path)
                print(f"✓ Validation report saved → {output_path}")
                return  # Success!

            except PermissionError as e:
                last_error = e
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 2  # Exponential backoff: 2s, 4s, 6s
                    print(f"  ⚠ File is locked (attempt {attempt + 1}/{max_retries}). "
                          f"Retrying in {wait_time} seconds...")
                    print(f"    (If you have the Excel file open, please close it)")
                    time.sleep(wait_time)
                else:
                    # Final attempt failed
                    print(f"\n✗ ERROR: Cannot save validation report - file is locked")
                    print(f"  File path: {output_path}")
                    print(f"  Reason: The file is open in Excel or locked by OneDrive")
                    print(f"\n  Solutions:")
                    print(f"    1. Close the Excel file if it's open")
                    print(f"    2. Close Excel completely and retry")
                    print(f"    3. Pause OneDrive sync temporarily")
                    print(f"    4. Check file permissions")
                    raise PermissionError(
                        f"Cannot save {output_path.name} after {max_retries} attempts. "
                        f"File is locked - please close Excel or check file permissions."
                    ) from last_error

            except Exception as e:
                # Other errors - don't retry
                print(f"✗ Error saving validation report: {e}")
                raise

    def get_error_count(self) -> int:
        """Get total number of errors"""
        return len(self.errors)

    def has_errors(self) -> bool:
        """Check if there are any errors"""
        return len(self.errors) > 0

    def get_errors_by_severity(self, severity: str) -> List[ValidationError]:
        """Get errors filtered by severity"""
        return [e for e in self.errors if e.severity == severity]

    def print_summary(self):
        """Print a text summary of validation errors"""
        if not self.errors:
            print("\n✓ No validation errors found")
            return

        print(f"\n✗ Found {len(self.errors)} validation error(s):")
        print("=" * 80)

        for idx, error in enumerate(self.errors[:10], 1):  # Show first 10
            location = f"Line {error.line_number}" if error.line_number else "Unknown location"
            print(f"{idx}. [{error.severity}] {error.xml_file} - {location}")
            print(f"   Type: {error.error_type}")
            print(f"   Description: {error.error_description}")
            print()

        if len(self.errors) > 10:
            print(f"... and {len(self.errors) - 10} more errors")
            print("See the Excel report for complete details")

        print("=" * 80)


# Convenience function for pipeline integration
def generate_validation_report(
    xml_file: str,
    dtd_error_log,
    reference_errors: List[str],
    output_path: Path,
    book_title: str = "Unknown"
) -> ValidationReportGenerator:
    """
    Generate a complete validation report from all error sources.

    Args:
        xml_file: Name of XML file being validated
        dtd_error_log: lxml DTD error log
        reference_errors: List of reference validation errors
        output_path: Where to save the Excel report
        book_title: Title of the book

    Returns:
        ValidationReportGenerator instance with all errors
    """
    generator = ValidationReportGenerator()

    # Add DTD errors
    if dtd_error_log:
        dtd_errors = generator.parse_dtd_errors(xml_file, dtd_error_log)
        for error in dtd_errors:
            generator.add_error(error)

    # Add reference errors
    if reference_errors:
        generator.add_reference_errors(xml_file, reference_errors)

    # Generate Excel report
    if generator.has_errors():
        generator.generate_excel_report(output_path, book_title)
        generator.print_summary()
    else:
        print(f"\n✓ No validation errors found for {book_title}")

    return generator
