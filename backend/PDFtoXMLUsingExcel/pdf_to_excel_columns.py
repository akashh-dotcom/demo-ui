import os
import subprocess
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import statistics
import argparse
import re
import statistics


# -------------------------------------------------------------
# Script Detection Configuration (Superscripts/Subscripts)
# -------------------------------------------------------------
# These thresholds detect tiny fragments (superscripts/subscripts) that should
# be merged with adjacent larger text. Very strict criteria to avoid false
# positives like drop caps or large first letters.

SCRIPT_MAX_WIDTH = 15               # Max width for scripts (drop caps are ~30-50px)
SCRIPT_MAX_HEIGHT = 14              # Max height for scripts (drop caps are ~36-48px, increased from 12 to catch 13px subscripts)
SCRIPT_MAX_TEXT_LENGTH = 3          # Max characters (usually 1-3)
SCRIPT_MAX_HORIZONTAL_GAP = 5       # Must be within 5px horizontally
SUPERSCRIPT_MIN_TOP_DIFF = -3       # Superscript can be 3px above parent
SUPERSCRIPT_MAX_TOP_DIFF = 3        # Or 3px below parent top
SUBSCRIPT_MIN_TOP_DIFF = 3          # Subscript is 3-10px below parent
SUBSCRIPT_MAX_TOP_DIFF = 10         # Maximum 10px below
SCRIPT_MAX_HEIGHT_RATIO = 0.75      # Script must be <75% of parent height

# Symbols to exclude from script detection (avoid false positives)
EXCLUDE_SYMBOLS = {'°', '™', '®', '©', '•', '·', '◦', '▪', '½', '¼', '¾', '⅓', '→', '←', '↑', '↓', '…', '‥'}


# -------------------------------------------------------------
# Script Detection Functions (Phase 1)
# -------------------------------------------------------------

def is_script_size(fragment):
    """
    Check if fragment meets size criteria for being a superscript/subscript.
    Very strict to avoid detecting drop caps or large first letters.
    """
    if fragment["width"] >= SCRIPT_MAX_WIDTH:
        return False
    if fragment["height"] >= SCRIPT_MAX_HEIGHT:
        return False
    
    text = fragment.get("text", "").strip()
    if len(text) > SCRIPT_MAX_TEXT_LENGTH:
        return False
    if not text:
        return False
    
    return True


def is_excluded_symbol(text):
    """Check if text is a symbol that should not be treated as script."""
    text = text.strip()
    
    if text in EXCLUDE_SYMBOLS:
        return True
    
    # Only allow alphanumeric scripts (excludes most symbols)
    if not text.replace('^', '').replace('_', '').isalnum():
        return True
    
    return False


def find_adjacent_parent(script_fragment, all_fragments, script_index):
    """
    Find the parent fragment for a potential superscript/subscript.
    
    Parent must be:
    - Larger in height than script
    - Adjacent horizontally (within 5px)
    - Close vertically (within 10px using TOP position)
    
    Returns (parent_index, parent_fragment) or None.
    """
    script_left = script_fragment["left"]
    script_right = script_left + script_fragment["width"]
    script_top = script_fragment["top"]
    script_height = script_fragment["height"]
    
    candidates = []
    
    for i, other in enumerate(all_fragments):
        if i == script_index:
            continue
        
        # Must be larger than script
        if other["height"] <= script_height:
            continue
        
        # Script must be significantly smaller (height ratio check)
        height_ratio = script_height / other["height"]
        if height_ratio >= SCRIPT_MAX_HEIGHT_RATIO:
            continue
        
        # Check horizontal adjacency
        other_left = other["left"]
        other_right = other_left + other["width"]
        
        # Is script to the right of other? (most common)
        gap_right = script_left - other_right
        if 0 <= gap_right <= SCRIPT_MAX_HORIZONTAL_GAP:
            # Check vertical proximity using TOP (not baseline!)
            top_diff = abs(script_top - other["top"])
            if top_diff <= SUBSCRIPT_MAX_TOP_DIFF:
                candidates.append((i, other, gap_right, top_diff))
        
        # Is script to the left of other? (rare)
        gap_left = other_left - script_right
        if 0 <= gap_left <= SCRIPT_MAX_HORIZONTAL_GAP:
            # Check vertical proximity using TOP (not baseline!)
            top_diff = abs(script_top - other["top"])
            if top_diff <= SUBSCRIPT_MAX_TOP_DIFF:
                candidates.append((i, other, gap_left, top_diff))
    
    if not candidates:
        return None
    
    # Choose closest candidate (smallest horizontal gap, then smallest vertical gap)
    candidates.sort(key=lambda x: (x[2], x[3]))
    
    parent_idx, parent, _, _ = candidates[0]
    return (parent_idx, parent)


def detect_script_type(script_fragment, parent_fragment):
    """
    Determine if script is superscript or subscript using TOP position.
    
    Key insight: Use TOP position, not baseline!
    Baseline = top + height is misleading for fragments with different heights.
    
    Returns "superscript", "subscript", or None.
    """
    # Calculate TOP difference (not baseline!)
    top_diff = script_fragment["top"] - parent_fragment["top"]
    
    # Superscript: within ±3px of parent top (but not >= 3px which is subscript)
    # Examples: 10⁷, x², aⁿ, references¹
    if SUPERSCRIPT_MIN_TOP_DIFF <= top_diff < SUBSCRIPT_MIN_TOP_DIFF:
        return "superscript"
    
    # Subscript: 3-10px below parent top
    # Examples: H₂O, B₀, a₁
    elif SUBSCRIPT_MIN_TOP_DIFF <= top_diff <= SUBSCRIPT_MAX_TOP_DIFF:
        return "subscript"
    
    return None


def detect_and_mark_scripts(fragments):
    """
    Phase 1: Detect and mark superscripts/subscripts using TOP position.
    
    This modifies fragments in-place by adding:
    - is_script: bool
    - script_type: "superscript" or "subscript"
    - script_parent_idx: index of parent fragment
    
    IMPORTANT: This does NOT change grouping logic!
    Baseline grouping remains unchanged, preserving drop caps and large letters.
    """
    # Add original index to each fragment
    for i, f in enumerate(fragments):
        f["original_idx"] = i
    
    # Detect scripts
    script_count = 0
    for i, f in enumerate(fragments):
        # Default: not a script
        f["is_script"] = False
        f["script_type"] = None
        f["script_parent_idx"] = None
        
        # Check size criteria
        if not is_script_size(f):
            continue
        
        # Check if excluded symbol
        text = f.get("text", "").strip()
        if is_excluded_symbol(text):
            continue
        
        # Find adjacent parent fragment
        parent_result = find_adjacent_parent(f, fragments, i)
        if not parent_result:
            continue
        
        parent_idx, parent = parent_result
        
        # Determine script type using TOP position (not baseline!)
        script_type = detect_script_type(f, parent)
        if not script_type:
            continue
        
        # Mark as script
        f["is_script"] = True
        f["script_type"] = script_type
        f["script_parent_idx"] = parent_idx
        script_count += 1
    
    return script_count


def merge_script_with_parent(parent, scripts):
    """
    Merge one or more scripts with their parent fragment.
    
    NOW TRACKS ORIGINAL FRAGMENTS for RittDocDTD-compliant output:
    - Stores original_fragments list including parent and scripts
    - Preserves script_type metadata for proper inline element generation
    
    Args:
        parent: Parent fragment
        scripts: List of script fragments to merge (sorted by position)
    
    Returns:
        Merged fragment
    """
    merged = dict(parent)  # Copy parent
    
    # Sort scripts by left position
    scripts = sorted(scripts, key=lambda s: s["left"])
    
    # NEW: Initialize fragment tracking
    if "original_fragments" in parent:
        # Parent already has tracking from previous merge
        merged["original_fragments"] = parent["original_fragments"].copy()
    else:
        # Start tracking with parent
        parent_copy = dict(parent)
        parent_copy.pop("original_fragments", None)
        merged["original_fragments"] = [parent_copy]
    
    # Merge text with caret (^) for superscripts, underscore (_) for subscripts
    merged_text = parent["text"]
    for script in scripts:
        script_text = script["text"]
        
        if script["script_type"] == "superscript":
            merged_text += "^" + script_text
        else:  # subscript
            merged_text += "_" + script_text
        
        # NEW: Track the script fragment
        script_copy = dict(script)
        script_copy.pop("original_fragments", None)
        merged["original_fragments"].append(script_copy)
    
    merged["text"] = merged_text
    merged["norm_text"] = " ".join(merged_text.split()).lower()
    
    # Merge inner_xml if present (preserve formatting)
    if "inner_xml" in parent:
        merged["inner_xml"] = parent.get("inner_xml", "")
        for script in scripts:
            merged["inner_xml"] += script.get("inner_xml", script["text"])
    
    # Expand bounding box to include all scripts
    for script in scripts:
        script_right = script["left"] + script["width"]
        merged_right = merged["left"] + merged["width"]
        if script_right > merged_right:
            merged["width"] = script_right - merged["left"]
        
        # Adjust height if script extends beyond
        script_bottom = script["top"] + script["height"]
        merged_bottom = merged["top"] + merged["height"]
        if script_bottom > merged_bottom:
            merged["height"] = script_bottom - merged["top"]
    
    # Mark as having merged scripts
    merged["has_merged_scripts"] = True
    merged["merged_script_count"] = len(scripts)
    
    return merged


def merge_scripts_across_rows(rows, all_fragments):
    """
    Phase 3: Merge scripts with their parents across rows.
    
    After baseline grouping, find scripts marked in Phase 1 and
    merge them with their parent fragments even if in different rows.
    
    This is the key to fixing superscript/subscript merging while
    preserving correct baseline grouping for drop caps and large letters.
    
    Args:
        rows: List of rows (each row is list of fragments)
        all_fragments: All fragments (for looking up by original_idx)
    
    Returns:
        Updated rows with scripts merged
    """
    # Build index: original_idx -> fragment
    frag_by_idx = {}
    for row in rows:
        for f in row:
            orig_idx = f.get("original_idx")
            if orig_idx is not None:
                frag_by_idx[orig_idx] = f
    
    # Find all scripts and group by parent
    scripts_by_parent = {}
    script_indices = set()
    
    for row in rows:
        for f in row:
            if f.get("is_script"):
                parent_idx = f.get("script_parent_idx")
                if parent_idx is not None:
                    if parent_idx not in scripts_by_parent:
                        scripts_by_parent[parent_idx] = []
                    scripts_by_parent[parent_idx].append(f)
                    script_indices.add(f.get("original_idx"))
    
    # Merge scripts into their parents
    merged_rows = []
    
    for row in rows:
        new_row = []
        
        for f in row:
            orig_idx = f.get("original_idx")
            
            # Skip if this fragment is a script (will be merged into parent)
            if orig_idx in script_indices:
                continue
            
            # Check if this fragment is a parent with scripts to merge
            if orig_idx in scripts_by_parent:
                scripts = scripts_by_parent[orig_idx]
                merged = merge_script_with_parent(f, scripts)
                new_row.append(merged)
            else:
                new_row.append(f)
        
        if new_row:
            merged_rows.append(new_row)
    
    return merged_rows


# -------------------------------------------------------------
# pdftohtml -xml runner
# -------------------------------------------------------------
def run_pdftohtml_xml(pdf_path, out_xml_path):
    """
    Run `pdftohtml -xml` to convert the PDF into an XML that we can parse.

    If out_xml_path is None, we create a .xml next to the PDF.
    """
    if out_xml_path is None:
        base, _ = os.path.splitext(pdf_path)
        out_xml_path = base + "_pdftohtml.xml"

    cmd = [
        "pdftohtml",
        "-xml",
        "-hidden",
        "-nodrm",
        "-i",
        "-enc",
        "UTF-8",
        pdf_path,
        out_xml_path,
    ]
    print("Running pdftohtml (this may take a few minutes for large PDFs)...")
    print("Command:", " ".join(cmd))
    
    try:
        # Run with a reasonable timeout (10 minutes for very large PDFs)
        result = subprocess.run(cmd, check=True, timeout=600, capture_output=True, text=True)
        print("✓ pdftohtml completed successfully")
        return out_xml_path
    except subprocess.TimeoutExpired:
        print("ERROR: pdftohtml timed out after 10 minutes")
        raise
    except subprocess.CalledProcessError as e:
        print(f"ERROR: pdftohtml failed with exit code {e.returncode}")
        if e.stderr:
            print(f"stderr: {e.stderr}")
        raise


# -------------------------------------------------------------
# Reading-order & line-grouping helpers
# -------------------------------------------------------------
def assign_reading_order_from_rows(fragments, rows):
    """
    Overwrite reading_order_index for all fragments using the
    already-constructed `rows` (baseline groups) and column IDs.

    - Each row gets a 'row_col' (dominant column for that row)
    - Rows are ordered column-major
    - Inside each row, fragments are ordered left-to-right
    """
    if not rows:
        return

    # Build row metadata
    row_infos = []
    for row in rows:
        if not row:
            continue
        # All frags in this row share the same row_index
        r_index = row[0]["row_index"]
        baseline = statistics.mean(f["baseline"] for f in row)
        # Dominant col_id in this row (ignore 0 unless everything is 0)
        col_ids = [f["col_id"] for f in row]
        non_zero = [c for c in col_ids if c > 0]
        if non_zero:
            row_col = statistics.mode(non_zero)
        else:
            row_col = 0

        row_infos.append(
            {
                "row_index": r_index,
                "baseline": baseline,
                "row_col": row_col,
                "fragments": row,
            }
        )

    if not row_infos:
        return

    # Which non-zero columns exist?
    col_ids = sorted({ri["row_col"] for ri in row_infos if ri["row_col"] > 0})

    # If there's no positive column, just do a simple top-to-bottom order
    if not col_ids:
        ordered_rows = sorted(row_infos, key=lambda r: r["baseline"])
    else:
        first_col = col_ids[0]
        first_col_min_y = min(
            r["baseline"] for r in row_infos if r["row_col"] == first_col
        )

        # 1) full-width rows above columns
        ordered_rows = sorted(
            [r for r in row_infos if r["row_col"] == 0 and r["baseline"] < first_col_min_y],
            key=lambda r: r["baseline"],
        )

        # 2) then columns in column-major order
        for c in col_ids:
            col_rows = [r for r in row_infos if r["row_col"] == c]
            col_rows_sorted = sorted(col_rows, key=lambda r: r["baseline"])
            ordered_rows.extend(col_rows_sorted)

        # 3) finally, any remaining full-width rows below columns
        remaining_full_width = [
            r
            for r in row_infos
            if r["row_col"] == 0 and r["baseline"] >= first_col_min_y
        ]
        remaining_full_width = sorted(remaining_full_width, key=lambda r: r["baseline"])
        ordered_rows.extend(remaining_full_width)

    # Now assign reading_order_index across all fragments in that row sequence
    ro_idx = 1
    for ri in ordered_rows:
        # Sort fragments left-to-right in the row
        row_frags = sorted(ri["fragments"], key=lambda f: f["left"])
        for f in row_frags:
            f["reading_order_index"] = ro_idx
            ro_idx += 1


def assign_reading_order_blocks(fragments, rows):
    """
    Assign reading_order_block to all fragments based on vertical position and col_id.

    Block assignment strategy (interleaved based on baseline):
      - Sort fragments by baseline (top to bottom reading order)
      - Increment block number whenever col_id changes
      - This naturally handles:
          * Full-width content above columns
          * Column 1, Column 2, etc.
          * Full-width content between columns
          * Full-width content below columns
    
    Examples:
      Title(0) → Col1(1) → Col2(2) → Footnote(0)  =  Blocks: 1, 2, 3, 4
      Title(0) → Col1(1) → Figure(0) → Col2(2)    =  Blocks: 1, 2, 3, 4
      Col1(1) → Col2(2) → Col3(3)                 =  Blocks: 1, 2, 3
    """
    if not fragments:
        return

    # Collect all unique col_ids
    all_col_ids = sorted({f["col_id"] for f in fragments if f["col_id"] is not None})

    # If everything is single column, assign Block 1 to all
    if len(all_col_ids) <= 1:
        for f in fragments:
            f["reading_order_block"] = 1
        return

    # Sort fragments by baseline (top to bottom), then by left position
    # This ensures we process fragments in natural reading order
    sorted_frags = sorted(fragments, key=lambda f: (f["baseline"], f["left"]))
    
    # Assign blocks based on col_id transitions
    # When col_id changes, we're moving to a new logical block
    block_num = 0
    prev_col_id = None
    
    for frag in sorted_frags:
        current_col_id = frag["col_id"]
        
        # Start a new block when col_id changes
        if current_col_id != prev_col_id:
            block_num += 1
            prev_col_id = current_col_id
        
        frag["reading_order_block"] = block_num


def compute_baseline_tolerance(baselines):
    """
    Compute how far apart two baselines can be and still be treated
    as the same row/line.
    """
    if len(baselines) < 2:
        return 2.0
    b_sorted = sorted(baselines)
    diffs = [
        b_sorted[i + 1] - b_sorted[i]
        for i in range(len(b_sorted) - 1)
        if b_sorted[i + 1] > b_sorted[i]
    ]
    if not diffs:
        return 2.0
    line_spacing = statistics.median(diffs)
    tol = min(2.0, line_spacing * 0.4)
    return tol


def group_fragments_into_lines(fragments, baseline_tol):
    """
    Given a list of fragments (already sorted by baseline, left),
    group them into rows based on baseline tolerance.
    """
    lines = []
    current = []
    current_baseline = None

    for f in fragments:
        b = f["baseline"]
        if current_baseline is None:
            current = [f]
            current_baseline = b
        elif abs(b - current_baseline) <= baseline_tol:
            current.append(f)
        else:
            lines.append(current)
            current = [f]
            current_baseline = b

    if current:
        lines.append(current)
    return lines

def merge_inline_fragments_in_row(row, gap_tolerance=1.5, space_width=1.0):
    """
    Merge adjacent fragments on the same baseline using a 3-phase rule
    with ± tolerance.
    
    NOW TRACKS ORIGINAL FRAGMENTS for RittDocDTD-compliant output:
    - Stores original_fragments list with all source fragments
    - Preserves font, size, and other metadata for each fragment
    - Enables output of inline elements (<phrase>, <emphasis>, etc.)

      Let:
        gap = next.left - (current.left + current.width)

      Phase 1) Trailing space detection:
         if current.text ends with " " AND next.text does NOT start with " ",
         then if |gap| <= gap_tolerance (≈ 0, so ±1 around 0 by default)
         → merge (space already present in current text).

      Phase 2) Inline-style split (no extra visible gap):
         if |gap| <= gap_tolerance   (≈ 0, so ±1 around 0 by default)
         → merge, regardless of next.text.

      Phase 3) Space-start continuation:
         if phases (1) and (2) fail AND next.text starts with " ",
         then if |gap - space_width| <= gap_tolerance
         (≈ 1, so ±1 around 1 by default)
         → merge.

      Otherwise, start a new logical fragment.

    gap_tolerance and space_width are in the same units as left/width
    (usually PDF points).
    """

    if not row:
        return []

    # Sort left-to-right
    row = sorted(row, key=lambda f: f["left"])

    merged = []
    current = dict(row[0])  # copy so we don't mutate original
    
    # NEW: Track original fragments for RittDocDTD compliance
    first_frag = dict(row[0])
    first_frag.pop("original_fragments", None)  # Prevent double-nesting
    current["original_fragments"] = [first_frag]

    for f in row[1:]:
        txt = f.get("text", "")
        current_txt = current.get("text", "")

        # Compute the horizontal gap between current and next
        base_end = current["left"] + current["width"]
        gap = f["left"] - base_end

        should_merge = False
        
        # --- SPECIAL CASE: Bullet point merging ---
        # Detect if current is a bullet character and next is text
        # Bullets are often positioned differently (different baseline/height)
        # So we need more lenient merging for bullets
        BULLET_CHARS = {'•', '●', '○', '■', '□', '▪', '▫', '·', '-', '*', '–', '—', '→', '⇒', '▸', '►'}
        current_stripped = current_txt.strip()
        
        if current_stripped in BULLET_CHARS and len(current_stripped) == 1:
            # Current is a bullet character - merge with following text if reasonably close
            # Allow larger gap (up to 20px) since bullets are often positioned differently
            if gap <= 20.0:  # More lenient for bullets
                should_merge = True

        # --- Phase 1: trailing space detection ---
        # If current ends with space and next does NOT start with space
        if not should_merge and current_txt.endswith(" ") and not txt.startswith(" "):
            # Check if gap is small (approximately zero)
            if abs(gap) <= gap_tolerance:
                should_merge = True

        # --- Phase 2: inline-style / no-gap merge ---
        if not should_merge:
            nogap = abs(gap) <= gap_tolerance
            if nogap:
                should_merge = True

        # --- Phase 3: starts-with-space + "space gap" (± tolerance) ---
        if not should_merge:
            if txt.startswith(" "):
                space_gap_ok = abs(gap - space_width) <= gap_tolerance
                if space_gap_ok:
                    should_merge = True

        if should_merge:
            # Merge: append text as-is (keep whatever spaces are in txt)
            current["text"] = current.get("text", "") + txt
            current["norm_text"] = " ".join(current["text"].split()).lower()

            # Merge XML content to preserve formatting
            current["inner_xml"] = current.get("inner_xml", "") + f.get("inner_xml", txt)

            # Expand width to cover the new fragment
            prev_end = current["left"] + current["width"]
            right = max(prev_end, f["left"] + f["width"])
            current["width"] = right - current["left"]
            
            # NEW: Track the merged fragment
            frag_copy = dict(f)
            frag_copy.pop("original_fragments", None)  # Prevent double-nesting
            current["original_fragments"].append(frag_copy)
        else:
            # Start a new logical fragment
            merged.append(current)
            current = dict(f)
            
            # NEW: Initialize tracking for new fragment
            frag_copy = dict(f)
            frag_copy.pop("original_fragments", None)
            current["original_fragments"] = [frag_copy]

    merged.append(current)
    return merged




# ---------------------------------------
# Fragment filtering (headers/footers)
# ---------------------------------------
def should_skip_fragment(norm_txt, top, height, page_height, seen_footer_texts):

    # 1) Skip if outside visible page render area
    if top > page_height * 1.05:     # below page
        return True
    if top < -20:                    # above page
        return True

    # 2) Skip file names, indesign junk, timestamps
    if re.search(r"\.indd\b", norm_txt):
        return True
    if re.search(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", norm_txt):
        return True  # dates like 12/18/18
    if re.search(r"\b\d{1,2}:\d{2}\s*(am|pm)\b", norm_txt):
        return True
    if re.fullmatch(r"[a-z0-9_\-]+\s+vi|vii|iv", norm_txt):
        return True

    # 3) Skip extremely small-height invisible text
    # (common for print-layer artifacts)
    if int(height) < 6: 
        return True

    # 4) Header/footer filtering - skip repeated text at page edges
    # Check if text is in header zone (top 12%) or footer zone (bottom 12%)
    if page_height > 0:
        norm_top = top / page_height
        is_header_zone = norm_top < 0.12
        is_footer_zone = norm_top > 0.85

        if (is_header_zone or is_footer_zone) and norm_txt in seen_footer_texts:
            return True

        # 5) Skip standalone page numbers at header/footer zones
        # These are captured as page IDs, don't need them in content
        if is_header_zone or is_footer_zone:
            text_stripped = norm_txt.strip()
            # Arabic page numbers (1-9999)
            if re.match(r'^\d{1,4}$', text_stripped):
                return True
            # Roman numerals (i, ii, iii, iv, v, vi, vii, viii, ix, x, etc.)
            if re.match(r'^[ivxlcdm]+$', text_stripped, re.IGNORECASE):
                return True

    return False


# -------------------------------------------------------------
# Column detection
# -------------------------------------------------------------
def detect_column_starts(fragments, page_width, max_cols=4,
                         min_cluster_size=15, min_cluster_ratio=0.10):
    """
    Detect approximate x-start of each text column by clustering fragment
    left positions. Then discard/merge tiny clusters (like vertical
    'INTRODUCTION' labels) into the nearest major column.

    CRITICAL FIX: Only consider an x-position as a column start if it has
    sufficient vertical distribution (multiple different baselines).
    This prevents line continuations from being treated as separate columns.

    - min_cluster_size: minimum number of fragments for a cluster
      to be considered a real column.
    - min_cluster_ratio: minimum fraction of all fragments on the page
      for a cluster to be considered a real column.
    """

    if not fragments:
        return []

    xs = sorted(f["left"] for f in fragments)
    if len(xs) == 1:
        return xs

    # --- Initial 1D clustering based on gaps in X ---
    # But also track which fragments belong to each cluster for vertical extent check
    clusters = []
    cluster_fragments = []  # Track fragments for each cluster
    current = [xs[0]]
    
    column_gap_threshold = page_width * 0.25

    # PERFORMANCE FIX: Build x_to_frags mapping more efficiently
    # Group fragments by their x position
    x_to_frags = {}
    for f in fragments:
        x = f["left"]
        if x not in x_to_frags:
            x_to_frags[x] = []
        x_to_frags[x].append(f)
    
    # Start with fragments at first x position
    current_frags = x_to_frags.get(xs[0], []).copy()

    for x in xs[1:]:
        mean_current = sum(current) / len(current)
        if abs(x - mean_current) <= column_gap_threshold:
            current.append(x)
            # Add all fragments at this x position
            if x in x_to_frags:
                current_frags.extend(x_to_frags[x])
        else:
            clusters.append(current)
            cluster_fragments.append(current_frags)
            current = [x]
            current_frags = x_to_frags.get(x, []).copy()
    clusters.append(current)
    cluster_fragments.append(current_frags)

    # --- NEW: Check vertical extent of each cluster ---
    # A true column should have fragments at many different vertical positions (baselines)
    # Line continuations and inline elements typically only appear on a few lines
    # Real columns have substantial vertical distribution (15-30+ lines of text)
    min_unique_baselines = 12  # A column should span at least 12 different lines
    baseline_tolerance = 2.0  # Group baselines within 2 pixels as same line

    valid_clusters = []
    valid_cluster_fragments = []

    for cluster, cluster_frags in zip(clusters, cluster_fragments):
        # Get unique baselines in this cluster (with tolerance)
        # PERFORMANCE FIX: Use a more efficient algorithm to count unique baseline groups
        baselines = sorted(set(f["baseline"] for f in cluster_frags))
        
        # Fast baseline grouping: sequential scan with tolerance
        unique_baseline_groups = []
        current_group_baseline = None
        
        for b in baselines:
            if current_group_baseline is None:
                # Start first group
                current_group_baseline = b
                unique_baseline_groups.append(b)
            elif abs(b - current_group_baseline) > baseline_tolerance:
                # Start new group (far enough from current group)
                current_group_baseline = b
                unique_baseline_groups.append(b)
            # else: within tolerance of current group, don't count as unique
        
        num_unique_lines = len(unique_baseline_groups)
        
        # Only keep clusters with sufficient vertical distribution
        if num_unique_lines >= min_unique_baselines:
            valid_clusters.append(cluster)
            valid_cluster_fragments.append(cluster_frags)
    
    # If no valid clusters found, treat entire page as single column
    if not valid_clusters:
        return [xs[0]]
    
    # Replace clusters with valid ones
    clusters = valid_clusters
    cluster_fragments = valid_cluster_fragments

    # --- If too many clusters, merge closest until ≤ max_cols ---
    while len(clusters) > max_cols:
        best_i, best_j, best_dist = None, None, None
        for i in range(len(clusters)):
            for j in range(i + 1, len(clusters)):
                mi = sum(clusters[i]) / len(clusters[i])
                mj = sum(clusters[j]) / len(clusters[j])
                dist = abs(mi - mj)
                if best_dist is None or dist < best_dist:
                    best_dist = dist
                    best_i, best_j = i, j
        if best_i is None:
            break
        merged = clusters[best_i] + clusters[best_j]
        merged_frags = cluster_fragments[best_i] + cluster_fragments[best_j]
        clusters = [
            c for k, c in enumerate(clusters)
            if k not in (best_i, best_j)
        ]
        cluster_fragments = [
            cf for k, cf in enumerate(cluster_fragments)
            if k not in (best_i, best_j)
        ]
        clusters.append(merged)
        cluster_fragments.append(merged_frags)

    # --- Filter out tiny "fake" columns (like vertical INTRODUCTION) ---
    total_points = sum(len(c) for c in clusters)
    if total_points == 0:
        return []

    major_clusters = []
    minor_clusters = []

    for c in clusters:
        size = len(c)
        if size >= min_cluster_size and size >= total_points * min_cluster_ratio:
            major_clusters.append(c)
        else:
            minor_clusters.append(c)

    # If everything is "minor", just use all clusters as-is
    if not major_clusters:
        major_clusters = clusters
        minor_clusters = []

    # Merge each minor cluster into the nearest major cluster by mean X
    if minor_clusters:
        major_means = [sum(c) / len(c) for c in major_clusters]
        for c in minor_clusters:
            m = sum(c) / len(c)
            nearest_idx = min(range(len(major_means)), key=lambda i: abs(m - major_means[i]))
            major_clusters[nearest_idx].extend(c)

    # Final column starts
    col_starts = [sum(c) / len(c) for c in major_clusters]
    return sorted(col_starts)


def is_vertical_spine_text(text, left, width, height,
                           page_width, page_height, rotation_deg):
    """
    Heuristic to filter out vertical spine text at the right border of the page,
    e.g. 'INTRODUCTION' printed vertically.

    We consider a fragment a vertical spine candidate if:
      - text is very short (<= 3 chars after stripping),
      - it sits near the right margin of the page,
      - and it is clearly vertical, either:
          * rotated ~90/270 degrees, OR
          * very tall & narrow.
    """
    if not text:
        return False

    t = text.strip()
    if len(t) > 3:
        return False

    # Near right border (tweak 0.8 if needed)
    if left <= page_width * 0.8:
        return False

    # Normal horizontal text: rotation 0 and not tall/narrow
    is_vertical_rot = abs(rotation_deg) in (90, 270)
    is_tall_narrow = height > width * 2.0

    if not (is_vertical_rot or is_tall_narrow):
        return False

    return True


def assign_column_ids(fragments, page_width, col_starts):
    """
    Assign a column id to each fragment on the page.

    col_id:
      1..N = one of the detected column clusters
      0    = full-width (spanning the entire page / multiple columns)
    """
    if not fragments:
        return

    # If there's effectively only one column start, treat everything as col 1
    if len(col_starts) <= 1:
        for f in fragments:
            f["col_id"] = 1
        return

     # Instead of using width/page_width, treat a fragment as "full-width"
    # only if it nearly touches BOTH left and right margins.
    margin_ratio = 0.05  # 5% of page width as margin
    left_margin  = page_width * margin_ratio
    right_margin = page_width * (1.0 - margin_ratio)

    # Calculate column boundaries (midpoints between adjacent column starts)
    # These boundaries define the territory for each column
    boundaries = []
    for i in range(len(col_starts) - 1):
        midpoint = (col_starts[i] + col_starts[i + 1]) / 2.0
        boundaries.append(midpoint)

    for f in fragments:
        left  = f["left"]
        right = f["left"] + f["width"]
        width = f["width"]

        # Full-width if it essentially spans from near-left to near-right
        if left <= left_margin and right >= right_margin:
            f["col_id"] = 0  # Fixed: full-width should be 0
        elif width >= page_width * 0.45:
            f["col_id"] = 0  # Fixed: wide content should be 0
        else:
            # OPTION 1: Assign based on LEFT edge position relative to boundaries
            # This handles indented text correctly - indented text in Col 1
            # still has its left edge in Col 1's territory
            if left < boundaries[0]:
                # Left edge is before first boundary → Column 1
                f["col_id"] = 1
            elif len(boundaries) > 1 and left >= boundaries[-1]:
                # Left edge is after last boundary → Last column
                f["col_id"] = len(col_starts)
            else:
                # Find which column territory the left edge falls into
                for i in range(len(boundaries)):
                    if i == len(boundaries) - 1:
                        # Between last boundary and end
                        f["col_id"] = i + 2
                        break
                    elif left < boundaries[i + 1]:
                        # Left edge is before the next boundary
                        f["col_id"] = i + 1
                        break


def reassign_misclassified_col0_fragments(fragments, page_width, col_starts):
    """
    Reassign ColID 0 fragments to the correct column if they're clearly
    within a single column's bounds and not actually full-width.

    This fixes cases where fragments were incorrectly classified as full-width (ColID 0)
    but are actually within a specific column.
    """
    if not fragments or len(col_starts) <= 1:
        return

    margin_ratio = 0.05
    left_margin = page_width * margin_ratio
    right_margin = page_width * (1.0 - margin_ratio)

    for f in fragments:
        # Only process fragments currently assigned to ColID 0
        if f["col_id"] != 0:
            continue

        left = f["left"]
        right = f["left"] + f["width"]
        width = f["width"]
        x_center = (left + right) / 2.0

        # Check if fragment actually spans full width - if so, keep ColID 0
        if left <= left_margin and right >= right_margin:
            continue  # Actually full-width, keep ColID 0

        # Check if fragment is very wide - if so, keep ColID 0
        if width >= page_width * 0.45:
            continue  # Wide enough to be full-width, keep ColID 0

        # Fragment is ColID 0 but doesn't span full width
        # Reassign to nearest column based on center position
        best = min(col_starts, key=lambda c: abs(x_center - c))
        f["col_id"] = col_starts.index(best) + 1


def group_col0_by_vertical_gap(fragments, typical_line_height, page_width=None, page_height=None):
    """
    Group ColID 0 fragments based on vertical gap - ONLY for wide fragments.

    When a fragment gets ColID 0 assigned, check the vertical gap to the next fragment.
    If the gap is small (within typical line height range) AND the next fragment is
    also wide enough to be full-width, assign ColID 0 to the next fragment.

    This ensures that multi-line titles, captions, and other full-width content
    that should logically be together all get ColID 0, WITHOUT incorrectly
    converting narrow column text to ColID 0.

    Special handling for footnote zone (bottom 25% of page):
    - Continuation lines can be narrow (e.g., "Accessed June, 2020.")
    - These should inherit parent's ColID 0 regardless of width

    Args:
        fragments: List of fragments (must have baseline, col_id, width, left)
        typical_line_height: Typical line height for gap threshold
        page_width: Page width for determining if fragment is truly full-width
        page_height: Page height for determining footnote zone
    """
    if not fragments:
        return

    # If we don't have page_width, skip this function to avoid incorrect ColID 0 propagation
    if page_width is None:
        return

    # Sort by baseline (top to bottom)
    sorted_frags = sorted(fragments, key=lambda f: f["baseline"])

    # Maximum vertical gap to consider fragments as part of the same ColID 0 group
    # Use 1.5x typical line height as threshold (same as paragraph detection)
    max_gap = typical_line_height * 1.5

    # Width thresholds for considering a fragment as "wide enough" for ColID 0
    # A fragment must be at least 40% of page width to be grouped as ColID 0
    min_width_ratio = 0.40
    min_width_for_col0 = page_width * min_width_ratio

    # Footnote zone threshold (bottom 25% of page)
    # In this zone, narrow continuation lines should inherit parent's ColID 0
    footnote_threshold = page_height * 0.75 if page_height else float('inf')

    i = 0
    while i < len(sorted_frags):
        current = sorted_frags[i]

        # Only process if current fragment has ColID 0
        if current["col_id"] != 0:
            i += 1
            continue

        # Check if current fragment is in footnote zone
        in_footnote_zone = current.get("top", 0) >= footnote_threshold

        # Current fragment has ColID 0, check subsequent fragments
        j = i + 1
        while j < len(sorted_frags):
            next_frag = sorted_frags[j]

            # Calculate vertical gap
            current_bottom = current["top"] + current["height"]
            next_top = next_frag["top"]
            vertical_gap = next_top - current_bottom

            # If gap is too large, stop grouping
            if vertical_gap > max_gap:
                break

            # Check if we should propagate ColID 0 to next fragment
            next_width = next_frag.get("width", 0)

            if next_frag["col_id"] != 0:
                # In footnote zone: propagate ColID 0 to continuation lines regardless of width
                # Outside footnote zone: only propagate if fragment is wide enough
                if in_footnote_zone or next_width >= min_width_for_col0:
                    # Fragment is wide enough OR we're in footnote zone - assign ColID 0
                    next_frag["col_id"] = 0
                    current = next_frag
                    # Update footnote zone status for the new current fragment
                    in_footnote_zone = current.get("top", 0) >= footnote_threshold
                else:
                    # Fragment is too narrow (within a column), stop propagating
                    break
            elif next_frag["col_id"] == 0:
                # Already ColID 0, continue
                current = next_frag
                # Update footnote zone status for the new current fragment
                in_footnote_zone = current.get("top", 0) >= footnote_threshold
            else:
                break

            j += 1

        # Move to the next ungrouped fragment
        i = j if j > i + 1 else i + 1


def maintain_col0_within_baseline(fragments, baseline_tol):
    """
    Maintain ColID 0 for fragments on the same baseline.

    If a fragment on a baseline has ColID 0, all subsequent fragments on the same
    baseline (within baseline_tol) should also get ColID 0 until we hit a line break.

    This fixes cases where the last small fragment on a ColID 0 line gets incorrectly
    assigned to ColID 1.
    """
    if not fragments:
        return

    # Sort by baseline, then left position
    sorted_frags = sorted(fragments, key=lambda f: (f["baseline"], f["left"]))

    # Group fragments by baseline
    baseline_groups = []
    current_group = []
    current_baseline = None

    for f in sorted_frags:
        b = f["baseline"]
        if current_baseline is None:
            current_group = [f]
            current_baseline = b
        elif abs(b - current_baseline) <= baseline_tol:
            current_group.append(f)
        else:
            if current_group:
                baseline_groups.append(current_group)
            current_group = [f]
            current_baseline = b

    if current_group:
        baseline_groups.append(current_group)

    # Process each baseline group
    for group in baseline_groups:
        # Check if ANY fragment in this baseline has ColID 0
        has_col0 = any(f["col_id"] == 0 for f in group)

        if has_col0:
            # If any fragment on this baseline is ColID 0, make ALL fragments ColID 0
            # This handles cases where the last small fragment gets misclassified
            for f in group:
                f["col_id"] = 0


def reclassify_footnote_rows_as_fullwidth(rows, page_width, page_height):
    """
    Detect and reclassify footnote rows that span multiple columns.

    Footnotes are often broken into multiple text fragments on the same baseline,
    where each fragment individually gets assigned to different columns, but
    collectively they form a full-width line at the bottom of the page.

    This function handles two cases:
    1. Rows with fragments from multiple columns that collectively span full-width
    2. Individual wide fragments (after merging) that should be reclassified as full-width

    This function detects such rows and reclassifies all fragments as col_id=0.
    """
    if not rows:
        return

    # Focus on bottom 25% of page where footnotes typically appear
    footnote_threshold = page_height * 0.75

    # Width threshold for considering a fragment as full-width
    width_threshold = page_width * 0.60  # 60% of page width

    for row in rows:
        if not row:
            continue

        # Check if row is in the footnote area (bottom 25%)
        row_top = min(f["top"] for f in row)
        if row_top < footnote_threshold:
            continue

        # Calculate collective span of all fragments in this row
        row_left = min(f["left"] for f in row)
        row_right = max(f["left"] + f["width"] for f in row)
        row_span = row_right - row_left

        # Get unique column IDs (excluding 0 which is already full-width)
        col_ids = {f["col_id"] for f in row if f["col_id"] != 0}

        # Case 1: Row has fragments from multiple columns spanning >75% page width
        # Case 2: Row collectively spans >75% page width (even if same column after merging)
        if row_span >= page_width * 0.75:
            # Reclassify all fragments in this row as full-width
            for f in row:
                f["col_id"] = 0
        # Case 3: Single wide fragment (after merging) that should be full-width
        elif len(row) == 1 and row[0]["width"] >= width_threshold and row[0]["col_id"] != 0:
            row[0]["col_id"] = 0


# -------------------------------------------------------------
# Single vs multi-column reading order
# -------------------------------------------------------------
def build_reading_order_single_column(fragments):
    """
    Fallback: if page is essentially single-column, reading order
    is just top-to-bottom, left-to-right.
    """
    sorted_frags = sorted(fragments, key=lambda f: (f["baseline"], f["left"]))
    idx = 1
    for f in sorted_frags:
        f["reading_order_index"] = idx
        idx += 1


def build_reading_order_multi_column(fragments):
    """
    For multi-column pages, sort by (col_id, baseline, left).
    """
    sorted_frags = sorted(
        fragments,
        key=lambda f: (0 if f["col_id"] is None else f["col_id"], f["baseline"], f["left"]),
    )
    idx = 1
    for f in sorted_frags:
        f["reading_order_index"] = idx
        idx += 1


# -------------------------------------------------------------
# Main PDF → Excel conversion
# -------------------------------------------------------------
def pdf_to_excel_with_columns(
    pdf_path,
    pdftohtml_xml_path=None,
    excel_output_path=None,
):
    # 1) Run pdftohtml -xml if needed
    if pdftohtml_xml_path is None:
        base, _ = os.path.splitext(pdf_path)
        pdftohtml_xml_path = base + "_pdftohtml.xml"

    if not os.path.exists(pdftohtml_xml_path):
        pdftohtml_xml_path = run_pdftohtml_xml(pdf_path, pdftohtml_xml_path)
    else:
        print(f"Using existing pdftohtml XML: {pdftohtml_xml_path}")

    # 1a) Backup original XML before we start processing
    base_xml, _ = os.path.splitext(pdftohtml_xml_path)
    backup_xml_path = base_xml + "_original.xml"
    if not os.path.exists(backup_xml_path):
        try:
            with open(pdftohtml_xml_path, "rb") as src, open(
                backup_xml_path, "wb"
            ) as dst:
                dst.write(src.read())
            print(f"Backed up original XML to: {backup_xml_path}")
        except Exception as e:
            print(f"Warning: could not back up XML: {e}")

    # 2) Parse XML
    tree = ET.parse(pdftohtml_xml_path)
    root = tree.getroot()

    # Prepare Excel workbook
    wb = Workbook()
    ws_ro = wb.active
    ws_ro.title = "ReadingOrder"

    ws_lines = wb.create_sheet("Lines")
    ws_img = wb.create_sheet("Images")
    ws_debug = wb.create_sheet("Debug")

    # Headers
    ws_ro.append(
        [
            "Page",
            "StreamIndex",
            "ReadingOrder",
            "ReadingOrderBlock",
            "ColID",
            "RowIndex",
            "Left",
            "Top",
            "Width",
            "Height",
            "Baseline",
            "Text",
        ]
    )

    ws_lines.append(
        [
            "Page",
            "RowIndex",
            "Baseline",
            "Col0_Text",
            "Col1_Text",
            "Col2_Text",
            "Col3_Text",
            "Col4_Text",
        ]
    )

    ws_img.append(["Page", "ImageIndex", "Left", "Top", "Width", "Height", "Label"])

    ws_debug.append(
        [
            "Page",
            "StreamIndex",
            "ColID",
            "RowIndex",
            "Baseline",
            "Left",
            "Top",
            "Width",
            "Height",
            "NormText",
        ]
    )

    seen_footer_texts = set()

    # Pre-pass: Scan pages to identify repeated header/footer text (chapter titles, page numbers)
    # Text that appears in same position on multiple pages is likely header/footer noise
    header_footer_candidates = {}  # {(norm_position, norm_text): count}
    page_elements_prescan = list(root.findall(".//page"))

    print(f"Pre-scanning {len(page_elements_prescan)} pages for header/footer patterns...")
    for page_elem in page_elements_prescan:
        page_height = float(page_elem.get("height", "0") or 0.0)
        page_width = float(page_elem.get("width", "0") or 0.0)

        if page_height <= 0:
            continue

        for t in page_elem.findall("text"):
            txt_raw = "".join(t.itertext())
            norm_txt = " ".join(txt_raw.split()).lower()
            
            # FIX 1: Add minimum text length requirement (5 chars)
            # Skip very short text (bullets, single chars) and very long text
            if not norm_txt or len(norm_txt) < 5 or len(norm_txt) > 100:
                continue
            
            # FIX 2: Exclude figure/table labels - they're legitimate content, not headers/footers
            if re.match(r'^(figure|table|fig\.?)\s+\d+', norm_txt, re.IGNORECASE):
                continue

            top = float(t.get("top", "0") or 0.0)
            left = float(t.get("left", "0") or 0.0)

            # Check if in header zone (top 12%) or footer zone (bottom 12%)
            norm_top = round(top / page_height, 2) if page_height > 0 else 0
            norm_left = round(left / page_width, 2) if page_width > 0 else 0

            is_header_zone = norm_top < 0.12
            is_footer_zone = norm_top > 0.88

            if is_header_zone or is_footer_zone:
                # Create a position key (rounded position + text)
                pos_key = (round(norm_top, 1), round(norm_left, 1), norm_txt)
                header_footer_candidates[pos_key] = header_footer_candidates.get(pos_key, 0) + 1

    # FIX 3: Increase minimum occurrence threshold for large documents
    # For large documents (1000+ pages), require higher repetition (at least 10 occurrences or 1% of pages)
    # For small documents, keep threshold at 3
    if len(page_elements_prescan) >= 500:
        min_occurrences = max(10, len(page_elements_prescan) // 100)  # At least 10, or 1% of pages
    else:
        min_occurrences = max(3, len(page_elements_prescan) // 10)  # At least 3, or 10% of pages
    
    print(f"  Using minimum occurrence threshold: {min_occurrences} (for {len(page_elements_prescan)} pages)")
    
    for (norm_top, norm_left, norm_txt), count in header_footer_candidates.items():
        if count >= min_occurrences:
            seen_footer_texts.add(norm_txt)
            print(f"  Header/footer pattern detected ({count}x): '{norm_txt[:50]}...' at position ({norm_top}, {norm_left})")

    if seen_footer_texts:
        print(f"  Total header/footer patterns to filter: {len(seen_footer_texts)}")
    else:
        print(f"  No repeated header/footer patterns detected")

    # Store all page data for return
    all_pages_data = {}

    # Iterate over pages
    page_elements = list(root.findall(".//page"))
    total_pages = len(page_elements)
    print(f"Processing {total_pages} pages...")
    
    for page_idx, page_elem in enumerate(page_elements, 1):
        page_number = int(page_elem.get("number", "0") or 0)
        
        # Progress indicator every 50 pages
        if page_idx % 50 == 0 or page_idx == 1:
            print(f"  Processing page {page_idx}/{total_pages} (page number: {page_number})")
        page_width = float(page_elem.get("width", "0") or 0.0)
        page_height = float(page_elem.get("height", "0") or 0.0)

        # Images → simple placeholders
        img_idx = 1
        for img in page_elem.findall("image"):
            l = float(img.get("left", "0") or 0.0)
            t = float(img.get("top", "0") or 0.0)
            w = float(img.get("width", "0") or 0.0)
            h = float(img.get("height", "0") or 0.0)
            ws_img.append([page_number, img_idx, l, t, w, h, "IMAGE_PLACEHOLDER"])
            img_idx += 1

        # Collect text fragments
        fragments = []
        page_number_fragments = []  # FIX 4: Separate list for page numbers (for page ID extraction)
        stream_index = 1

        for t in page_elem.findall("text"):
            # capture inner <b>, <i> etc.
            txt_raw = "".join(t.itertext())   # Get plain text for display/filtering
            txt = txt_raw                     # no strip
            if not txt:
                continue

            # Preserve inner XML structure for formatting (stores as XML string)
            inner_xml = ET.tostring(t, encoding="unicode", method="xml")
            # Remove the outer <text...> wrapper, keeping only inner content
            # Extract content between opening and closing <text> tags
            inner_content = inner_xml
            if inner_xml.startswith("<text"):
                start = inner_xml.find(">") + 1
                end = inner_xml.rfind("</text>")
                if start > 0 and end > start:
                    inner_content = inner_xml[start:end]
                elif inner_xml.endswith("/>"):
                    # Self-closing tag, no content
                    inner_content = txt_raw
                else:
                    inner_content = txt_raw
            else:
                inner_content = txt_raw

            left = float(t.get("left", "0") or 0.0)
            top = float(t.get("top", "0") or 0.0)
            width = float(t.get("width", "0") or 0.0)
            height = float(t.get("height", "0") or 0.0)
            baseline = top + height

            # Try a few common attribute names for rotation; fall back to 0 if missing
            rot_raw = t.get("rotation") or t.get("rotate") or t.get("rot") or "0"
            try:
                rotation_deg = int(float(rot_raw))
            except ValueError:
                rotation_deg = 0

            # Check for vertical spine text to skip
            # Filter out vertical spine text at page border (INTRODUCTION, etc.)
            if is_vertical_spine_text(txt, left, width, height,
                              page_width, page_height, rotation_deg):
                continue

            norm_txt = " ".join(txt.split()).lower()
            
            # FIX 4: Check if this is a standalone page number BEFORE filtering
            # Preserve page numbers for page ID extraction even if they'd be filtered from content
            is_page_number = False
            if page_height > 0:
                norm_top = top / page_height
                is_header_zone = norm_top < 0.12
                is_footer_zone = norm_top > 0.85
                
                if is_header_zone or is_footer_zone:
                    text_stripped = norm_txt.strip()
                    # Check for arabic numbers (1-9999) or roman numerals
                    if re.match(r'^\d{1,4}$', text_stripped) or re.match(r'^[ivxlcdm]+$', text_stripped, re.IGNORECASE):
                        is_page_number = True
                        # Store in separate list for page ID extraction
                        page_number_fragments.append({
                            "text": txt,
                            "norm_text": norm_txt,
                            "left": left,
                            "top": top,
                            "width": width,
                            "height": height,
                        })
            
            # Apply normal filtering (page numbers will be filtered from main content but preserved above)
            if should_skip_fragment(norm_txt, top, height, page_height, seen_footer_texts):
                continue

            fragments.append({
                "stream_index": stream_index,
                "text": txt,
                "inner_xml": inner_content,  # Preserve formatting tags
                "norm_text": norm_txt,
                "left": left,
                "top": top,
                "width": width,
                "height": height,
                "baseline": baseline,
                "col_id": None,
                "row_index": None,
                "reading_order_index": None,
                "reading_order_block": None,
            })
            stream_index += 1

        if not fragments:
            continue
        
        # Debug: Log if page numbers were found
        if page_number_fragments:
            print(f"  Page {page_number}: Found {len(page_number_fragments)} page number(s) for ID extraction")

        # Sort by baseline & left for line grouping
        fragments.sort(key=lambda f: (f["baseline"], f["left"]))
        
        # ===== Phase 1: Detect superscripts/subscripts =====
        # Detect and mark scripts BEFORE grouping into rows.
        # Uses TOP position (not baseline) to find scripts adjacent to larger text.
        # Very strict criteria (w<15, h<12) to avoid false positives (drop caps, etc.)
        script_count = detect_and_mark_scripts(fragments)
        if script_count > 0:
            print(f"  Page {page_number}: Detected {script_count} superscript(s)/subscript(s)")

        # Warn if page has many fragments (potential performance issue)
        if len(fragments) > 1000:
            print(f"  Page {page_number}: {len(fragments)} fragments (large page, may take longer)")

        # Column detection for this page (initially still uses all fragments;
        # merge of inline fragments will run before final row/col assignments)
        col_starts = detect_column_starts(fragments, page_width, max_cols=4)
        assign_column_ids(fragments, page_width, col_starts)

        # Reassign misclassified ColID 0 fragments to correct columns
        reassign_misclassified_col0_fragments(fragments, page_width, col_starts)

        # Maintain ColID 0 for all fragments on the same baseline
        # (fixes issue where last small fragment gets assigned to wrong column)
        baselines_for_col0 = [f["baseline"] for f in fragments]
        baseline_tol_for_col0 = compute_baseline_tolerance(baselines_for_col0) if baselines_for_col0 else 2.0
        maintain_col0_within_baseline(fragments, baseline_tol_for_col0)

        # Decide reading order: single vs multi column
        """
        unique_cols = {f["col_id"] for f in fragments}
        if len([c for c in unique_cols if c > 0]) <= 1:
            build_reading_order_single_column(fragments)
        else:
            build_reading_order_multi_column(fragments)

        # Group into lines based on baseline
        baselines = [f["baseline"] for f in fragments]
        baseline_tol = compute_baseline_tolerance(baselines)
        rows = group_fragments_into_lines(fragments, baseline_tol)
        """

        # (1) First pass: group into rows and merge inline fragments within each row
        baselines = [f["baseline"] for f in fragments]
        baseline_tol = compute_baseline_tolerance(baselines)
        raw_rows = group_fragments_into_lines(fragments, baseline_tol)
        
        # ===== Phase 3: Merge scripts across rows =====
        # Merge superscripts/subscripts with their parents even if in different rows.
        # This fixes cases like "10^7" where "7" is in a different row due to baseline difference.
        raw_rows = merge_scripts_across_rows(raw_rows, fragments)

        merged_fragments = []
        for row in raw_rows:
            merged_fragments.extend(merge_inline_fragments_in_row(row))

        fragments = merged_fragments
        if not fragments:
            continue

        # Re-sort after merging
        fragments.sort(key=lambda f: (f["baseline"], f["left"]))

        # (2) Now group again into rows with merged fragments
        baselines = [f["baseline"] for f in fragments]
        baseline_tol = compute_baseline_tolerance(baselines)
        rows = group_fragments_into_lines(fragments, baseline_tol)

        # (2a) Reclassify footnote rows that span multiple columns as full-width
        reclassify_footnote_rows_as_fullwidth(rows, page_width, page_height)

        # (2b) Group ColID 0 fragments by vertical gap - only for wide fragments
        # Calculate typical line height for gap detection
        line_heights = [f["height"] for f in fragments if f["height"] > 0]
        typical_line_height = sorted(line_heights)[len(line_heights) // 2] if line_heights else 12.0
        group_col0_by_vertical_gap(fragments, typical_line_height, page_width=page_width, page_height=page_height)

        # (3) assign row indices
        row_idx = 1
        for row in rows:
            for f in row:
                f["row_index"] = row_idx
            row_idx += 1

        # (4) Overwrite reading-order using the new row-based method
        assign_reading_order_from_rows(fragments, rows)

        # (5) Assign reading-order blocks
        assign_reading_order_blocks(fragments, rows)

        # Store page data for return (copy fragments to preserve all info)
        all_pages_data[page_number] = {
            "page_width": page_width,
            "page_height": page_height,
            "fragments": [dict(f) for f in fragments],  # deep copy
            "page_number_fragments": page_number_fragments,  # FIX 4: Preserve page numbers for page ID extraction
        }

        # ---------------------------------------------------------
        # Excel output: ReadingOrder + Debug
        # ---------------------------------------------------------
        for f in fragments:
            ws_ro.append(
                [
                    page_number,
                    f["stream_index"],
                    f["reading_order_index"],
                    f["reading_order_block"],
                    f["col_id"],
                    f["row_index"],
                    f["left"],
                    f["top"],
                    f["width"],
                    f["height"],
                    f["baseline"],
                    f["text"],
                ]
            )
            ws_debug.append(
                [
                    page_number,
                    f["stream_index"],
                    f["col_id"],
                    f["row_index"],
                    f["baseline"],
                    f["left"],
                    f["top"],
                    f["width"],
                    f["height"],
                    f["norm_text"],
                ]
            )

        # ---------------------------------------------------------
        # Excel output: Lines sheet (grouped by row & col)
        # ---------------------------------------------------------
        for row in rows:
            if not row:
                continue

            y_top = min(f["top"] for f in row)
            b_row = statistics.mean(f["baseline"] for f in row)
            row_index = row[0]["row_index"]

            # col_id -> list of frags in that row
            by_col = {}
            for f in row:
                c = f["col_id"]
                by_col.setdefault(c, []).append(f)

            col_ids = sorted(by_col.keys())
            col_count = len(col_ids)

            # up to 5 buckets: 0,1,2,3,4
            texts = ["", "", "", "", ""]
            for c in col_ids:
                seg_frags = sorted(by_col[c], key=lambda f: f["left"])
                seg_text = " ".join(f["text"] for f in seg_frags)
                idx_c = c if 0 <= c <= 4 else 4
                # Append with separator if multiple segments for same col
                if texts[idx_c]:
                    texts[idx_c] += " | " + seg_text
                else:
                    texts[idx_c] = seg_text

            ws_lines.append(
                [
                    page_number,
                    row_index,
                    b_row,
                    texts[0],
                    texts[1],
                    texts[2],
                    texts[3],
                    texts[4],
                ]
            )

    # ---------------------------------------------------------
    # Save Excel
    # ---------------------------------------------------------
    print(f"\nCompleted processing all {total_pages} pages")
    print("Saving Excel file...")
    
    if excel_output_path is None:
        base, _ = os.path.splitext(pdf_path)
        excel_output_path = base + "_columns.xlsx"

    wb.save(excel_output_path)
    print(f"✓ Excel saved to: {excel_output_path}")

    # Return structured data for unified XML generation
    return {
        "excel_path": excel_output_path,
        "pdftohtml_xml_path": pdftohtml_xml_path,
        "pages": all_pages_data,
    }


# -------------------------------------------------------------
# CLI
# -------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert PDF (via pdftohtml -xml) into Excel with column-aware reading order."
    )
    parser.add_argument("pdf_path", help="Path to input PDF")
    parser.add_argument("--xml", dest="pdftohtml_xml_path", help="Output XML path")
    parser.add_argument("--excel", dest="excel_output_path", help="Output Excel path")

    args = parser.parse_args()

    pdf_to_excel_with_columns(
        pdf_path=args.pdf_path,
        pdftohtml_xml_path=args.pdftohtml_xml_path,
        excel_output_path=args.excel_output_path,
    )
